import calendar
import re
from zipfile import ZipFile
from datetime import date
from io import BytesIO

import dateparser
import openpyxl
from beancount.core import data
from beancount.core.data import Note, Transaction

from ..accounts import accounts
from . import (get_account_by_guess,
               get_income_account_by_guess, replace_flag)
from .base import Base
from .deduplicate import Deduplicate

Account零钱通 = 'Assets:Company:WeChat:Lingqiantong'
Account收入红包 = 'Income:RedBag'
Account支出红包 = 'Expenses:Social:RedBag'
Account余额 = 'Assets:Balances:WeChat'

class WeChat(Base):

    def __init__(self, filename, byte_content, entries, option_map):
        if re.search(r'微信支付账单.*\.zip$', filename):
            password = input('微信账单密码：')
            z = ZipFile(BytesIO(byte_content), 'r')
            z.setpassword(bytes(password, 'utf-8'))
            filelist = z.namelist()
            if len(filelist) == 2 and re.search(r'微信支付.*\.xlsx$', filelist[1]):
                byte_content = z.read(filelist[1])
        self.workbook = openpyxl.load_workbook(BytesIO(byte_content))
        self.worksheet = self.workbook.active
        first_cell = self.worksheet.cell(row=1, column=1).value
        if first_cell != '微信支付账单明细':
            raise Exception('Not WeChat Trade Record!')

        print('Import WeChat: ' + str(self.worksheet.cell(row=3, column=1).value))
        self.deduplicate = Deduplicate(entries, option_map, self.__class__.__name__)

    def parse(self):
        worksheet = self.worksheet

        # 找到表头行（通常在第17行）
        header_row = 17
        headers = []
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=header_row, column=col).value
            if cell_value:
                headers.append(cell_value.strip())
            else:
                break

        transactions = []

        # 从第18行开始读取数据
        for row_num in range(header_row + 1, worksheet.max_row + 1):
            row_data = {}
            has_data = False

            # 读取每一行的数据
            for col_idx, header in enumerate(headers):
                cell_value = worksheet.cell(row=row_num, column=col_idx + 1).value
                if cell_value is not None:
                    row_data[header] = str(cell_value).strip()
                    has_data = True
                else:
                    row_data[header] = ''

            # 如果行为空，跳过
            if not has_data or not row_data.get('商品'):
                continue

            print("Importing {} at {}".format(row_data['商品'], row_data['交易时间']))

            meta = {}
            time = dateparser.parse(row_data['交易时间'])
            meta['wechat_trade_no'] = row_data['交易单号']
            meta['trade_time'] = row_data['交易时间']
            meta['timestamp'] = str(time.timestamp()).replace('.0', '')
            account = get_account_by_guess(row_data['交易对方'], row_data['商品'], time)
            flag = "*"
            amount_string = row_data['金额(元)'].replace('¥', '')
            amount = float(amount_string)

            if row_data['商户单号'] != '/':
                meta['shop_trade_no'] = row_data['商户单号']

            if row_data['备注'] != '/':
                meta['note'] = row_data['备注']

            meta = data.new_metadata(
                'beancount/core/testing.beancount',
                12345,
                meta
            )

            tags = []
            if row_data['商品'] == '亲属卡':
                tags.append("love-pay")
            if len(tags) == 0:
                tags = data.EMPTY_SET

            entry = Transaction(
                meta,
                date(time.year, time.month, time.day),
                '*',
                row_data['交易对方'],
                row_data['商品'],
                tags,
                data.EMPTY_SET, []
            )

            status = row_data['当前状态']

            if status == '支付成功' or status == '朋友已收钱' or status == '已全额退款' or '已退款' in status or status == '已转账' or status == '充值成功':
                if '转入零钱通' in row_data['交易类型']:
                    entry = entry._replace(payee='')
                    entry = entry._replace(narration='转入零钱通')
                    data.create_simple_posting(
                        entry, Account零钱通, amount_string, 'CNY')
                else:
                    if '微信红包' in row_data['交易类型']:
                        account = Account支出红包
                        if entry.narration == '/':
                            entry = entry._replace(narration=row_data['交易类型'])
                    else:
                        account = get_account_by_guess(
                            row_data['交易对方'], row_data['商品'], time)
                    # if account == "Unknown":
                    #	entry = replace_flag(entry, '!')
                    if (status == '已全额退款' or '已退款' in status) and row_data['收/支'] == '收入':
                        amount_string = '-' + amount_string
                    data.create_simple_posting(
                        entry, account, amount_string, 'CNY')
                data.create_simple_posting(
                    entry, accounts[row_data['支付方式']], None, None)
            elif row_data['当前状态'] == '已存入零钱' or row_data['当前状态'] == '已收钱' or row_data['当前状态'] == '提现已到账':
                if '微信红包' in row_data['交易类型']:
                    if entry.narration == '/':
                        entry = entry._replace(narration=row_data['交易类型'])
                    data.create_simple_posting(entry, Account收入红包, None, 'CNY')
                else:
                    income = get_income_account_by_guess(
                        row_data['交易对方'], row_data['商品'], time)
                    if income == 'Income:Unknown':
                        entry = replace_flag(entry, '!')
                    data.create_simple_posting(entry, income, None, 'CNY')
                data.create_simple_posting(
                    entry, Account余额, amount_string, 'CNY')
            else:
                print('Unknown row', row_data)

            #b = printer.format_entry(entry)
            # print(b)
            if not self.deduplicate.find_duplicate(entry, amount, 'wechat_trade_no'):
                transactions.append(entry)

        self.deduplicate.apply_beans()
        return transactions
